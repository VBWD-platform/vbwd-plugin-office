"""``.csv``/``.xlsx`` import and export for VBWD Spreadsheets (S147-4
requirement #7).

**The non-negotiable bit, honoured by construction rather than by a special
case:** an unmapped Excel function is never detected up front and dropped —
the imported formula's SOURCE TEXT is stored verbatim, exactly like any
other formula, and handed to the engine's own ``recalculate``. The engine
already returns ``#NAME?`` for a function name that is not in its registry
(``engine._evaluate_function_call``) — so an unmapped function surfaces
automatically as "literal source plus ``#NAME?``" with no bespoke detection
logic to keep in sync with the registry (DRY: one place decides what a
function name resolves to). ``import_workbook`` (in
``sheet_editor_service.py``) builds the import report by scanning the
recalculation's own output for ``#NAME?`` cells — nothing here needs to know
which functions exist.

``.xlsx`` support is optional at the dependency level: if ``openpyxl`` is not
installed, :func:`import_xlsx`/:func:`export_xlsx` raise
:class:`OfficeSheetUnavailableFormatError` rather than silently degrading —
per the sprint's explicit escape hatch ("ship .csv only, report .xlsx as not
done" rather than a lossy but silent xlsx path).
"""
from __future__ import annotations

import csv
import datetime
import io
from dataclasses import dataclass
from typing import Optional

from plugins.office.office.sheet.cell import CellAddress, format_cell_reference
from plugins.office.office.sheet.engine import Cell, Workbook
from plugins.office.office.sheet.values import (
    BLANK,
    Blank,
    CellValue,
    coerce_to_text,
    is_error,
)
from plugins.office.office.services.exceptions import OfficeSheetUnavailableFormatError
from plugins.office.office.services.sheet_content import (
    DEFAULT_SHEET_NAME,
    SheetCeiling,
    parse_reference_within_ceiling,
)

CSV_FORMAT = "csv"
XLSX_FORMAT = "xlsx"
SUPPORTED_IMPORT_FORMATS = (CSV_FORMAT, XLSX_FORMAT)
SUPPORTED_EXPORT_FORMATS = (CSV_FORMAT, XLSX_FORMAT)

#: openpyxl truncates sheet titles at 31 characters — enforced on export so
#: a long Doc/Sheet name never raises deep inside the library call.
XLSX_MAX_SHEET_TITLE_LENGTH = 31

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME_TYPE = "text/csv"


@dataclass(frozen=True)
class ImportReportEntry:
    """One formula the import could not map to a registered function —
    still imported (source text preserved), never dropped."""

    sheet: str
    address: str
    formula: str

    def to_dict(self) -> dict:
        return {"sheet": self.sheet, "address": self.address, "formula": self.formula}


@dataclass(frozen=True)
class ImportedWorkbook:
    workbook: Workbook


def _literal_from_csv_text(text: str) -> CellValue:
    stripped = text.strip()
    if stripped == "":
        return BLANK
    try:
        return float(stripped)
    except ValueError:
        return text


def import_csv(
    data: bytes, ceiling: SheetCeiling, sheet_name: str = DEFAULT_SHEET_NAME
) -> ImportedWorkbook:
    """CSV carries literal values only — never a formula (the format has no
    notion of one)."""
    text = data.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))

    workbook = Workbook()
    sheet = workbook.add_sheet(sheet_name)
    for row_index, row in enumerate(rows, start=1):
        for column_index, raw_cell in enumerate(row, start=1):
            reference = format_cell_reference(
                CellAddress(sheet=None, column=column_index, row=row_index),
                include_sheet=False,
            )
            address = parse_reference_within_ceiling(reference, sheet_name, ceiling)
            value = _literal_from_csv_text(raw_cell)
            if isinstance(value, Blank):
                continue
            sheet.set_cell(address.column, address.row, Cell(value=value))
    return ImportedWorkbook(workbook=workbook)


def export_csv(workbook: Workbook, sheet_name: Optional[str] = None) -> bytes:
    """Export the CURRENT (caller-recalculated) values of one sheet as CSV —
    formulas do not survive a CSV round trip, only their computed values do."""
    name = (
        sheet_name
        if sheet_name in workbook.sheets
        else next(iter(workbook.sheets), DEFAULT_SHEET_NAME)
    )
    sheet = workbook.sheets.get(name)
    output = io.StringIO()
    writer = csv.writer(output)
    if sheet is not None and sheet.cells:
        max_row = max(row for _column, row in sheet.cells)
        max_column = max(column for column, _row in sheet.cells)
        for row in range(1, max_row + 1):
            writer.writerow(
                [
                    coerce_to_text(sheet.get_cell(column, row).value)
                    for column in range(1, max_column + 1)
                ]
            )
    return output.getvalue().encode("utf-8")


def _xlsx_literal(raw_value) -> CellValue:
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    if isinstance(raw_value, (datetime.datetime, datetime.date)):
        return raw_value
    return str(raw_value)


def import_xlsx(data: bytes, ceiling: SheetCeiling) -> ImportedWorkbook:
    try:
        import openpyxl
    except ImportError as import_error:
        raise OfficeSheetUnavailableFormatError(
            "openpyxl is not installed; .xlsx import is unavailable"
        ) from import_error

    source_workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    workbook = Workbook()
    for worksheet in source_workbook.worksheets:
        sheet = workbook.add_sheet(worksheet.title)
        for row in worksheet.iter_rows():
            for source_cell in row:
                if source_cell.value is None:
                    continue
                reference = format_cell_reference(
                    CellAddress(
                        sheet=None, column=source_cell.column, row=source_cell.row
                    ),
                    include_sheet=False,
                )
                address = parse_reference_within_ceiling(
                    reference, worksheet.title, ceiling
                )
                if isinstance(source_cell.value, str) and source_cell.value.startswith(
                    "="
                ):
                    sheet.set_cell(
                        address.column, address.row, Cell(formula=source_cell.value)
                    )
                else:
                    sheet.set_cell(
                        address.column,
                        address.row,
                        Cell(value=_xlsx_literal(source_cell.value)),
                    )
    return ImportedWorkbook(workbook=workbook)


def _xlsx_export_value(value: CellValue):
    if is_error(value):
        return str(value)
    if isinstance(value, Blank):
        return None
    return value


def export_xlsx(workbook: Workbook) -> bytes:
    try:
        import openpyxl
    except ImportError as import_error:
        raise OfficeSheetUnavailableFormatError(
            "openpyxl is not installed; .xlsx export is unavailable"
        ) from import_error

    output_workbook = openpyxl.Workbook()
    output_workbook.remove(output_workbook.active)
    for name, sheet in workbook.sheets.items():
        worksheet = output_workbook.create_sheet(
            title=name[:XLSX_MAX_SHEET_TITLE_LENGTH]
        )
        for (column, row), cell in sheet.cells.items():
            if cell.formula is not None:
                formula_text = (
                    cell.formula if cell.formula.startswith("=") else f"={cell.formula}"
                )
                worksheet.cell(row=row, column=column, value=formula_text)
            elif not isinstance(cell.value, Blank):
                worksheet.cell(
                    row=row, column=column, value=_xlsx_export_value(cell.value)
                )

    output = io.BytesIO()
    output_workbook.save(output)
    return output.getvalue()
